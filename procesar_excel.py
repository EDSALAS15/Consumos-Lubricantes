"""
procesar_excel.py
Convierte Consumosdeaceites.xlsx → datos.json
Ejecutado automáticamente por GitHub Actions cada vez que el Excel cambia.
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

EXCEL_FILE = 'Consumosdeaceites.xlsx'
JSON_FILE  = 'datos.json'

print(f"Leyendo {EXCEL_FILE}...")
wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
ws = wb.active

# Leer encabezados
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"Columnas: {headers}")

# Campos requeridos
CAMPOS = {
    'DIVISION':           'DIVISION',
    'ECONOMICO':          'ECONOMICO',
    'TIPO':               'TIPO',
    'MARCA DE UNIDAD':    'MARCA DE UNIDAD',
    'TIPO DE CARROCERIA': 'TIPO DE CARROCERIA',
    'MOTOR':              'MOTOR',
    'CAPACIDAD LTS':      'CAPACIDAD LTS',
    'FECHA CONSUMO':      'FECHA CONSUMO',
    'CANTIDAD':           'CANTIDAD',
}

# Índices de columnas
col_idx = {}
for campo in CAMPOS:
    for i, h in enumerate(headers):
        if h and str(h).strip().upper() == campo.upper():
            col_idx[campo] = i
            break
    if campo not in col_idx:
        print(f"  ADVERTENCIA: columna '{campo}' no encontrada")

print(f"Columnas encontradas: {col_idx}")

# ── Leer filas y construir estructuras optimizadas ───────────────
units_map  = {}
unit_list  = []
fechas_set = set()
pivot_tmp  = {}

print("Procesando filas...")
filas = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    filas += 1

    def get(campo):
        i = col_idx.get(campo)
        return row[i] if i is not None else None

    # División
    div_raw = get('DIVISION')
    div = str(div_raw).strip() if div_raw else ''
    if div.upper() == 'VICTORIA':
        div = 'VICTORIA'

    eco   = str(get('ECONOMICO') or '').replace('.0','').strip()
    tipo  = str(get('TIPO') or '').strip()
    marca = str(get('MARCA DE UNIDAD') or '').strip()
    carr  = str(get('TIPO DE CARROCERIA') or '').strip()
    motor = str(get('MOTOR') or '').strip()
    cap   = get('CAPACIDAD LTS')
    cap   = str(cap) if cap is not None else ''

    # Fecha — puede ser date, datetime o string
    fecha_raw = get('FECHA CONSUMO')
    if fecha_raw is None:
        continue
    if hasattr(fecha_raw, 'strftime'):
        fecha = fecha_raw.strftime('%Y-%m-%d')
    else:
        fecha = str(fecha_raw).strip()[:10]
    if not fecha or len(fecha) < 10:
        continue

    # Cantidad
    cant_raw = get('CANTIDAD')
    try:
        cant = float(cant_raw or 0)
    except (ValueError, TypeError):
        continue
    if cant == 0:
        continue

    fechas_set.add(fecha)

    unit_key = f"{div}|{eco}|{tipo}|{marca}|{carr}|{motor}|{cap}"
    if unit_key not in units_map:
        units_map[unit_key] = len(unit_list)
        unit_list.append([div, eco, tipo, marca, carr, motor, cap])

    pkey = unit_key + '||' + fecha
    pivot_tmp[pkey] = pivot_tmp.get(pkey, 0) + cant

print(f"  {filas} filas leídas")
print(f"  {len(unit_list)} unidades únicas")
print(f"  {len(fechas_set)} fechas únicas")
print(f"  {len(pivot_tmp)} celdas pivot")

wb.close()

# ── Vocabularios ─────────────────────────────────────────────────
da, ta, ma, ca, moa, caa = set(), set(), set(), set(), set(), set()
for u in unit_list:
    da.add(u[0]); ta.add(u[2]); ma.add(u[3])
    ca.add(u[4]); moa.add(u[5]); caa.add(u[6])

darr  = sorted(da)
tarr  = sorted(ta)
marr  = sorted(ma)
carr2 = sorted(ca)
moarr = sorted(moa)
caarr = sorted(caa)

INV = [darr, tarr, marr, carr2, moarr, caarr]

di   = {v:i for i,v in enumerate(darr)}
ti   = {v:i for i,v in enumerate(tarr)}
mi   = {v:i for i,v in enumerate(marr)}
ci   = {v:i for i,v in enumerate(carr2)}
moi  = {v:i for i,v in enumerate(moarr)}
cai  = {v:i for i,v in enumerate(caarr)}

FECHAS = sorted(fechas_set)
fi_map = {f:i for i,f in enumerate(FECHAS)}
nF     = len(FECHAS)

UNITS = [
    [di[u[0]], u[1], ti[u[2]], mi[u[3]], ci[u[4]], moi[u[5]], cai[u[6]]]
    for u in unit_list
]

# Pivot como [[ui, fi, qty*10], ...]
PIVOT = []
for pkey, qty in pivot_tmp.items():
    sep = pkey.rfind('||')
    uk  = pkey[:sep]
    f   = pkey[sep+2:]
    ui  = units_map[uk]
    fi  = fi_map.get(f)
    if fi is None:
        continue
    PIVOT.append([ui, fi, round(qty * 10)])

# ── Guardar JSON ─────────────────────────────────────────────────
output = {
    'INV':    INV,
    'UNITS':  UNITS,
    'FECHAS': FECHAS,
    'PIVOT':  PIVOT,
}

SEP = (',', ':')
json_str = json.dumps(output, separators=SEP, ensure_ascii=False)

with open(JSON_FILE, 'w', encoding='utf-8') as f:
    f.write(json_str)

size_kb = len(json_str.encode()) / 1024
print(f"\n✓ {JSON_FILE} generado: {size_kb:.0f} KB")
print(f"  INV vocabularios: {[len(v) for v in INV]}")
print(f"  UNITS: {len(UNITS)}")
print(f"  FECHAS: {len(FECHAS)} ({FECHAS[0]} → {FECHAS[-1]})")
print(f"  PIVOT: {len(PIVOT)} celdas")
